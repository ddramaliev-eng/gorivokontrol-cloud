# -*- coding: utf-8 -*-
"""
GorivoKontrol Cloud - уеб версия (Flask) на GorivoKontrol v2.3.

Пренесена бизнес логика от desktop (Tkinter) приложението, вкл. поправките от
v2.3 (коректно конвертиране лв/€ при сумиране на общи суми, защита срещу
намаляващ километраж, динамичен списък с видове гориво).

Данните се пазят в SQLite файл в DATA_DIR (по подразбиране ./data - в Railway
се задава DATA_DIR=/app/data и се монтира Volume на същия път, за да не се
губят данните при redeploy).
"""
import io
import os
import secrets
import sqlite3
from datetime import date, datetime, timedelta
from functools import wraps

from flask import (
    Flask, g, request, redirect, url_for, render_template,
    session, flash, send_file, abort, jsonify
)
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ---------------- Конфигурация ----------------
APP_PASSWORD = os.environ.get("APP_PASSWORD", "gorivo123")
SECRET_KEY = os.environ.get("SECRET_KEY") or secrets.token_hex(32)
DATA_DIR = os.environ.get("DATA_DIR", os.path.join(os.path.dirname(os.path.abspath(__file__)), "data"))
DB_PATH = os.path.join(DATA_DIR, "data.db")

DEFAULT_FUEL_TYPES = ["Дизел", "Бензин", "Газ", "Електро"]

# Валутен преход: след 31.12.2025 цените са в евро (официално въвеждане на
# еврото в България от 01.01.2026)
EUR_TRANSITION_DATE = datetime(2025, 12, 31)
EUR_BGN_FIXED_RATE = 1.95583
EUR_PERIOD_THRESHOLD = (EUR_TRANSITION_DATE + timedelta(days=1)).strftime("%Y")

app = Flask(__name__)
app.secret_key = SECRET_KEY


# ---------------- Бизнес логика (пренесена от v2.3) ----------------
def get_currency_for_date(date_iso: str) -> str:
    try:
        date_obj = datetime.strptime(date_iso, "%Y-%m-%d")
        return "€" if date_obj > EUR_TRANSITION_DATE else "лв"
    except Exception:
        return "лв"


def period_is_eur(period_label: str) -> bool:
    """period_label е 'YYYY' или 'YYYY-MM'; работи благодарение на
    лексикографското сравнение на низове, тъй като преходът е точно на
    01.01 - началото на година/месец."""
    return period_label >= EUR_PERIOD_THRESHOLD


def to_bgn(amount: float, period_label: str) -> float:
    return amount * EUR_BGN_FIXED_RATE if period_is_eur(period_label) else amount


def to_bgn_by_date(amount: float, date_iso: str) -> float:
    return amount * EUR_BGN_FIXED_RATE if get_currency_for_date(date_iso) == "€" else amount


def sum_prices_bgn(labels, prices) -> float:
    return sum(to_bgn(p, lbl) for lbl, p in zip(labels, prices))


def periods_are_mixed_currency(labels) -> bool:
    has_eur = any(period_is_eur(l) for l in labels)
    has_lv = any(not period_is_eur(l) for l in labels)
    return has_eur and has_lv


def display_date(iso: str) -> str:
    try:
        return datetime.strptime(iso, "%Y-%m-%d").strftime("%d.%m.%Y")
    except Exception:
        return iso or ""


def validate_positive_float(value, field_name: str) -> float:
    try:
        num = float(value)
    except (TypeError, ValueError):
        raise ValueError(f"Невалидна стойност за {field_name}")
    if num < 0:
        raise ValueError(f"{field_name} не може да бъде отрицателно число")
    return num


def validate_date_not_future(date_iso: str) -> str:
    try:
        d = date.fromisoformat(date_iso)
    except Exception:
        raise ValueError("Невалиден формат на дата.")
    if d > date.today():
        raise ValueError("Датата не може да бъде в бъдещето")
    return date_iso


def sanitize_filename(filename: str) -> str:
    dangerous = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
    for ch in dangerous:
        filename = filename.replace(ch, '_')
    return filename


# ---------------- База данни ----------------
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(exc=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    os.makedirs(DATA_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS vehicles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            registration TEXT NOT NULL UNIQUE,
            model TEXT,
            year INTEGER,
            fuel_type TEXT,
            start_odometer REAL DEFAULT 0
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS refuels (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vehicle_id INTEGER,
            date TEXT,
            odometer REAL,
            liters REAL,
            price_per_liter REAL,
            total_price REAL,
            note TEXT,
            FOREIGN KEY(vehicle_id) REFERENCES vehicles(id)
        )
    """)
    conn.commit()
    conn.close()


def get_fuel_type_options(db):
    rows = db.execute(
        "SELECT DISTINCT fuel_type FROM vehicles WHERE fuel_type IS NOT NULL AND TRIM(fuel_type)<>''"
    ).fetchall()
    db_types = [r[0] for r in rows]
    return sorted(set(DEFAULT_FUEL_TYPES) | set(db_types))


def get_max_odometer(db, vehicle_id, exclude_refuel_id=None):
    if exclude_refuel_id is not None:
        row = db.execute(
            "SELECT MAX(odometer) FROM refuels WHERE vehicle_id=? AND id<>?",
            (vehicle_id, exclude_refuel_id),
        ).fetchone()
    else:
        row = db.execute(
            "SELECT MAX(odometer) FROM refuels WHERE vehicle_id=?", (vehicle_id,)
        ).fetchone()
    return row[0] if row and row[0] is not None else None


# ---------------- Автентикация ----------------
def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("authenticated"):
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)
    return wrapped


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        pw = request.form.get("password", "")
        if secrets.compare_digest(pw, APP_PASSWORD):
            session["authenticated"] = True
            nxt = request.args.get("next") or url_for("home")
            return redirect(nxt)
        flash("Грешна парола.", "danger")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ---------------- Начало ----------------
@app.route("/")
@login_required
def home():
    db = get_db()
    total_vehicles = db.execute("SELECT COUNT(*) FROM vehicles").fetchone()[0]
    total_refuels = db.execute("SELECT COUNT(*) FROM refuels").fetchone()[0]
    fuel_stats = db.execute(
        "SELECT fuel_type, COUNT(*) FROM vehicles GROUP BY fuel_type"
    ).fetchall()
    return render_template(
        "home.html",
        total_vehicles=total_vehicles,
        total_refuels=total_refuels,
        fuel_stats=fuel_stats,
        eur_rate=EUR_BGN_FIXED_RATE,
    )


# ---------------- Автомобили ----------------
@app.route("/vehicles")
@login_required
def vehicles_list():
    db = get_db()
    vehicles = db.execute(
        "SELECT id, registration, model, year, fuel_type, start_odometer FROM vehicles ORDER BY registration"
    ).fetchall()
    return render_template("vehicles.html", vehicles=vehicles)


def _vehicle_form_values(form):
    reg = (form.get("registration") or "").strip()
    model = (form.get("model") or "").strip()
    year_str = (form.get("year") or "").strip()
    fuel = (form.get("fuel_type") or "").strip()
    start_km_str = (form.get("start_odometer") or "").strip()
    if not reg:
        raise ValueError("Регистрационният номер е задължителен.")
    year_val = None
    if year_str:
        try:
            year_val = int(year_str)
        except ValueError:
            raise ValueError("Годината трябва да бъде цяло число.")
    start_km_val = 0.0
    if start_km_str:
        start_km_val = validate_positive_float(start_km_str, "Началните километри")
    return reg, model, year_val, fuel, start_km_val


@app.route("/vehicles/add", methods=["GET", "POST"])
@login_required
def vehicle_add():
    if request.method == "POST":
        try:
            reg, model, year_val, fuel, start_km_val = _vehicle_form_values(request.form)
            db = get_db()
            db.execute(
                "INSERT INTO vehicles (registration, model, year, fuel_type, start_odometer) VALUES (?, ?, ?, ?, ?)",
                (reg, model, year_val, fuel, start_km_val),
            )
            db.commit()
            flash("Автомобилът е добавен.", "success")
            return redirect(url_for("vehicles_list"))
        except sqlite3.IntegrityError:
            flash("Вече съществува автомобил с този регистрационен номер.", "danger")
        except ValueError as e:
            flash(str(e), "danger")
    return render_template("vehicle_form.html", vehicle=None, fuel_types=DEFAULT_FUEL_TYPES, title="Добави автомобил")


@app.route("/vehicles/<int:vid>/edit", methods=["GET", "POST"])
@login_required
def vehicle_edit(vid):
    db = get_db()
    vehicle = db.execute("SELECT * FROM vehicles WHERE id=?", (vid,)).fetchone()
    if not vehicle:
        abort(404)
    if request.method == "POST":
        try:
            reg, model, year_val, fuel, start_km_val = _vehicle_form_values(request.form)
            db.execute(
                "UPDATE vehicles SET registration=?, model=?, year=?, fuel_type=?, start_odometer=? WHERE id=?",
                (reg, model, year_val, fuel, start_km_val, vid),
            )
            db.commit()
            flash("Промените са запазени.", "success")
            return redirect(url_for("vehicles_list"))
        except sqlite3.IntegrityError:
            flash("Регистрационният номер вече се използва.", "danger")
        except ValueError as e:
            flash(str(e), "danger")
    return render_template("vehicle_form.html", vehicle=vehicle, fuel_types=DEFAULT_FUEL_TYPES, title="Редактирай автомобил")


@app.route("/vehicles/<int:vid>/delete", methods=["POST"])
@login_required
def vehicle_delete(vid):
    db = get_db()
    db.execute("DELETE FROM refuels WHERE vehicle_id=?", (vid,))
    db.execute("DELETE FROM vehicles WHERE id=?", (vid,))
    db.commit()
    flash("Автомобилът и зарежданията му са изтрити.", "success")
    return redirect(url_for("vehicles_list"))


# ---------------- Зареждания ----------------
@app.route("/refuels")
@login_required
def refuels_list():
    db = get_db()
    vehicle_filter = request.args.get("vehicle", "")
    search = request.args.get("q", "").strip()
    vehicles = db.execute("SELECT id, registration FROM vehicles ORDER BY registration").fetchall()

    if vehicle_filter:
        rows = db.execute(
            """SELECT r.id, v.registration, r.date, r.odometer, r.liters, r.price_per_liter, r.total_price, r.note
               FROM refuels r JOIN vehicles v ON r.vehicle_id=v.id
               WHERE v.registration=? ORDER BY r.date""",
            (vehicle_filter,),
        ).fetchall()
    else:
        rows = db.execute(
            """SELECT r.id, v.registration, r.date, r.odometer, r.liters, r.price_per_liter, r.total_price, r.note
               FROM refuels r JOIN vehicles v ON r.vehicle_id=v.id
               ORDER BY r.date"""
        ).fetchall()

    items = []
    for r in rows:
        disp_date = display_date(r["date"])
        if search and (search not in disp_date and search not in str(r["odometer"])):
            continue
        items.append({
            "id": r["id"], "registration": r["registration"], "date": disp_date,
            "odometer": r["odometer"], "liters": r["liters"], "price": r["price_per_liter"],
            "total": r["total_price"], "note": r["note"],
            "currency": get_currency_for_date(r["date"]),
        })

    return render_template(
        "refuels.html", items=items, vehicles=vehicles,
        vehicle_filter=vehicle_filter, search=search,
    )


@app.route("/refuels/add", methods=["GET", "POST"])
@login_required
def refuel_add():
    db = get_db()
    vehicles = db.execute("SELECT id, registration FROM vehicles ORDER BY registration").fetchall()
    if not vehicles:
        flash("Няма добавени автомобили. Добавете автомобил първо.", "warning")
        return redirect(url_for("vehicle_add"))

    warning = None
    form_values = {"date": date.today().isoformat()}

    if request.method == "POST":
        form_values = dict(request.form)
        try:
            vid = int(request.form.get("vehicle_id"))
            date_iso = request.form.get("date", "")
            validate_date_not_future(date_iso)
            odo = validate_positive_float(request.form.get("odometer"), "Километри")
            liters = validate_positive_float(request.form.get("liters"), "Литри")
            price = validate_positive_float(request.form.get("price"), "Цена")
            note = (request.form.get("note") or "").strip()

            max_odo = get_max_odometer(db, vid)
            confirmed = request.form.get("confirm") == "1"
            if max_odo is not None and odo < max_odo and not confirmed:
                warning = (
                    f"Въведеният километраж ({odo:.0f} км) е по-малък от последния "
                    f"записан за този автомобил ({max_odo:.0f} км). Това ще доведе до "
                    "неверни изчисления на изминати км/разход."
                )
            else:
                total = liters * price
                db.execute(
                    """INSERT INTO refuels (vehicle_id, date, odometer, liters, price_per_liter, total_price, note)
                       VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (vid, date_iso, odo, liters, price, total, note),
                )
                db.commit()
                flash("Записът е добавен.", "success")
                return redirect(url_for("refuels_list"))
        except ValueError as e:
            flash(str(e), "danger")

    return render_template(
        "refuel_form.html", vehicles=vehicles, refuel=None, values=form_values,
        title="Добави зареждане", warning=warning,
    )


@app.route("/refuels/<int:rid>/edit", methods=["GET", "POST"])
@login_required
def refuel_edit(rid):
    db = get_db()
    refuel = db.execute("SELECT * FROM refuels WHERE id=?", (rid,)).fetchone()
    if not refuel:
        abort(404)
    vehicles = db.execute("SELECT id, registration FROM vehicles ORDER BY registration").fetchall()

    warning = None
    form_values = {
        "vehicle_id": refuel["vehicle_id"], "date": refuel["date"],
        "odometer": refuel["odometer"], "liters": refuel["liters"],
        "price": refuel["price_per_liter"], "note": refuel["note"],
    }

    if request.method == "POST":
        form_values = dict(request.form)
        try:
            vid = int(request.form.get("vehicle_id"))
            date_iso = request.form.get("date", "")
            validate_date_not_future(date_iso)
            odo = validate_positive_float(request.form.get("odometer"), "Километри")
            liters = validate_positive_float(request.form.get("liters"), "Литри")
            price = validate_positive_float(request.form.get("price"), "Цена")
            note = (request.form.get("note") or "").strip()

            max_odo = get_max_odometer(db, vid, exclude_refuel_id=rid)
            confirmed = request.form.get("confirm") == "1"
            if max_odo is not None and odo < max_odo and not confirmed:
                warning = (
                    f"Въведеният километраж ({odo:.0f} км) е по-малък от последния "
                    f"записан за този автомобил ({max_odo:.0f} км). Това ще доведе до "
                    "неверни изчисления на изминати км/разход."
                )
            else:
                total = liters * price
                db.execute(
                    """UPDATE refuels SET vehicle_id=?, date=?, odometer=?, liters=?, price_per_liter=?, total_price=?, note=?
                       WHERE id=?""",
                    (vid, date_iso, odo, liters, price, total, note, rid),
                )
                db.commit()
                flash("Промените са запазени.", "success")
                return redirect(url_for("refuels_list"))
        except ValueError as e:
            flash(str(e), "danger")

    return render_template(
        "refuel_form.html", vehicles=vehicles, refuel=refuel, values=form_values,
        title="Редактирай зареждане", warning=warning,
    )


@app.route("/refuels/<int:rid>/delete", methods=["POST"])
@login_required
def refuel_delete(rid):
    db = get_db()
    db.execute("DELETE FROM refuels WHERE id=?", (rid,))
    db.commit()
    flash("Записът е изтрит.", "success")
    return redirect(url_for("refuels_list"))


# ---------------- Отчети ----------------
def _build_report_data(db, vehicle, fuel_filter, grouping):
    """Връща (labels, liters_list, prices_list, kms_list, consumptions_list,
    currencies_list) - идентична логика на generate_report() от v2.3,
    поправена за коректно валутно сумиране."""
    if vehicle:
        query = """SELECT r.date, r.odometer, r.liters, r.total_price, v.fuel_type
                   FROM refuels r JOIN vehicles v ON r.vehicle_id=v.id
                   WHERE v.registration=?"""
        params = [vehicle]
        if fuel_filter:
            query += " AND v.fuel_type=?"
            params.append(fuel_filter)
        query += " ORDER BY r.date"
        rows = db.execute(query, params).fetchall()
    else:
        query = """SELECT v.registration, r.date, r.odometer, r.liters, r.total_price, v.fuel_type
                   FROM refuels r JOIN vehicles v ON r.vehicle_id=v.id"""
        params = []
        if fuel_filter:
            query += " WHERE v.fuel_type=?"
            params.append(fuel_filter)
        query += " ORDER BY v.registration, r.date"
        rows = db.execute(query, params).fetchall()

    groups = {}
    if vehicle:
        for row in rows:
            key = row["date"][:7] if grouping == "month" else row["date"][:4]
            g_ = groups.setdefault(key, {"liters": 0.0, "price": 0.0, "start_odo": None, "end_odo": None})
            g_["liters"] += row["liters"]
            g_["price"] += row["total_price"]
            if g_["start_odo"] is None:
                g_["start_odo"] = row["odometer"]
            g_["end_odo"] = row["odometer"]
    else:
        # авто-парк: км се смятат по превозно средство, после се сумират за периода
        veh_bucket = {}
        for row in rows:
            key = row["date"][:7] if grouping == "month" else row["date"][:4]
            g_ = groups.setdefault(key, {"liters": 0.0, "price": 0.0})
            g_["liters"] += row["liters"]
            g_["price"] += row["total_price"]
            veh_bucket.setdefault(key, {}).setdefault(row["registration"], []).append(
                (row["date"], row["odometer"])
            )

    labels = sorted(groups.keys())
    liters = [groups[k]["liters"] for k in labels]
    prices = [groups[k]["price"] for k in labels]
    kms, consumptions, currencies = [], [], []

    for k in labels:
        if vehicle:
            g_ = groups[k]
            km = (g_["end_odo"] - g_["start_odo"]) if (g_["start_odo"] is not None and g_["end_odo"] is not None) else 0
        else:
            km = 0
            for reg, data in veh_bucket.get(k, {}).items():
                if len(data) > 1:
                    data.sort()
                    km += data[-1][1] - data[0][1]
        kms.append(km)
        liters_k = groups[k]["liters"]
        consumptions.append((liters_k / km * 100) if km > 0 else 0)
        currencies.append("€" if period_is_eur(k) else "лв")

    return labels, liters, prices, kms, consumptions, currencies


@app.route("/reports")
@login_required
def reports():
    db = get_db()
    vehicles = [r[0] for r in db.execute("SELECT registration FROM vehicles ORDER BY registration").fetchall()]
    fuel_types = get_fuel_type_options(db)

    vehicle = request.args.get("vehicle", "")
    fuel_filter = request.args.get("fuel", "")
    grouping = request.args.get("grouping", "month")
    if grouping not in ("month", "year"):
        grouping = "month"

    report = None
    if vehicle or request.args.get("run") == "1":
        labels, liters, prices, kms, consumptions, currencies = _build_report_data(db, vehicle, fuel_filter, grouping)
        if labels:
            total_liters = sum(liters)
            total_price_bgn = sum_prices_bgn(labels, prices)
            total_km = sum(kms)
            avg100 = (total_liters / total_km * 100) if total_km > 0 else 0
            report = {
                "labels": labels, "liters": liters, "prices": prices,
                "kms": kms, "consumptions": consumptions, "currencies": currencies,
                "total_liters": total_liters, "total_price_bgn": total_price_bgn,
                "total_km": total_km, "avg100": avg100,
                "mixed_currency": periods_are_mixed_currency(labels),
            }

    return render_template(
        "reports.html", vehicles=vehicles, fuel_types=fuel_types,
        vehicle=vehicle, fuel_filter=fuel_filter, grouping=grouping,
        report=report, eur_rate=EUR_BGN_FIXED_RATE,
    )


def _autosize_and_style_header(ws, headers):
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15


@app.route("/reports/export.xlsx")
@login_required
def reports_export():
    db = get_db()
    vehicle = request.args.get("vehicle", "")
    fuel_filter = request.args.get("fuel", "")
    grouping = request.args.get("grouping", "month")
    labels, liters, prices, kms, consumptions, currencies = _build_report_data(db, vehicle, fuel_filter, grouping)
    if not labels:
        flash("Няма данни за експорт.", "warning")
        return redirect(url_for("reports", vehicle=vehicle, fuel=fuel_filter, grouping=grouping, run="1"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Обобщен отчет"
    headers = ["Период", "Литри", "Платено", "Валута", "Изминати км", "Л/100км"]
    _autosize_and_style_header(ws, headers)
    thin = Side(border_style="thin", color="000000")
    for row_i, lbl in enumerate(labels, 2):
        ws.cell(row=row_i, column=1, value=lbl)
        ws.cell(row=row_i, column=2, value=liters[row_i - 2])
        ws.cell(row=row_i, column=3, value=prices[row_i - 2])
        ws.cell(row=row_i, column=4, value=currencies[row_i - 2])
        ws.cell(row=row_i, column=5, value=kms[row_i - 2])
        ws.cell(row=row_i, column=6, value=consumptions[row_i - 2])
        for col in range(1, 7):
            cell = ws.cell(row=row_i, column=col)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(horizontal="center")
            if col in (2, 3, 5, 6):
                cell.number_format = "#,##0.00"

    total_row = len(labels) + 3
    total_liters = sum(liters)
    total_price_bgn = sum_prices_bgn(labels, prices)
    total_km = sum(kms)
    ws.cell(row=total_row, column=1, value="ОБЩО (лв)")
    ws.cell(row=total_row, column=2, value=total_liters)
    ws.cell(row=total_row, column=3, value=total_price_bgn)
    ws.cell(row=total_row, column=4, value="лв")
    ws.cell(row=total_row, column=5, value=total_km)
    ws.cell(row=total_row, column=6, value=(total_liters / total_km * 100) if total_km > 0 else 0)
    for col in range(1, 7):
        cell = ws.cell(row=total_row, column=col)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        if col in (2, 3, 5, 6):
            cell.number_format = "#,##0.00"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    today = date.today().strftime("%d%m%Y")
    name = f"Отчет_{vehicle or (fuel_filter or 'автопарк')}_{today}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=sanitize_filename(name),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/reports/detailed")
@login_required
def reports_detailed():
    db = get_db()
    vehicles = [r[0] for r in db.execute("SELECT registration FROM vehicles ORDER BY registration").fetchall()]
    fuel_types = get_fuel_type_options(db)
    vehicle = request.args.get("vehicle", "")
    fuel_filter = request.args.get("fuel", "")
    search = request.args.get("q", "").strip()

    query = """SELECT r.date, v.registration, v.fuel_type, r.odometer, r.liters,
                      r.price_per_liter, r.total_price, r.note
               FROM refuels r JOIN vehicles v ON r.vehicle_id=v.id WHERE 1=1"""
    params = []
    if vehicle:
        query += " AND v.registration=?"
        params.append(vehicle)
    if fuel_filter:
        query += " AND v.fuel_type=?"
        params.append(fuel_filter)
    query += " ORDER BY r.date"
    rows = db.execute(query, params).fetchall()

    items = []
    for r in rows:
        disp = display_date(r["date"])
        if search and (search not in disp and search not in str(r["odometer"]) and search not in r["registration"]):
            continue
        items.append({
            "date": disp, "registration": r["registration"], "fuel_type": r["fuel_type"],
            "odometer": r["odometer"], "liters": r["liters"], "price": r["price_per_liter"],
            "total": r["total_price"], "note": r["note"],
            "currency": get_currency_for_date(r["date"]),
        })

    return render_template(
        "detailed.html", items=items, vehicles=vehicles, fuel_types=fuel_types,
        vehicle=vehicle, fuel_filter=fuel_filter, search=search,
    )


@app.route("/reports/detailed/export.xlsx")
@login_required
def reports_detailed_export():
    db = get_db()
    vehicle = request.args.get("vehicle", "")
    fuel_filter = request.args.get("fuel", "")

    query = """SELECT r.date, v.registration, v.fuel_type, r.odometer, r.liters,
                      r.price_per_liter, r.total_price, r.note
               FROM refuels r JOIN vehicles v ON r.vehicle_id=v.id WHERE 1=1"""
    params = []
    if vehicle:
        query += " AND v.registration=?"
        params.append(vehicle)
    if fuel_filter:
        query += " AND v.fuel_type=?"
        params.append(fuel_filter)
    query += " ORDER BY r.date"
    rows = db.execute(query, params).fetchall()
    if not rows:
        flash("Няма данни за експорт.", "warning")
        return redirect(url_for("reports_detailed", vehicle=vehicle, fuel=fuel_filter))

    wb = Workbook()
    ws = wb.active
    ws.title = "Подробен отчет"
    headers = ["Дата", "Автомобил", "Гориво", "Км", "Литри", "Цена/л", "Общо", "Валута", "Бележка"]
    _autosize_and_style_header(ws, headers)
    thin = Side(border_style="thin", color="000000")
    total_liters = 0.0
    total_price_bgn = 0.0
    for row_i, r in enumerate(rows, 2):
        currency = get_currency_for_date(r["date"])
        ws.cell(row=row_i, column=1, value=display_date(r["date"]))
        ws.cell(row=row_i, column=2, value=r["registration"])
        ws.cell(row=row_i, column=3, value=r["fuel_type"])
        ws.cell(row=row_i, column=4, value=r["odometer"])
        ws.cell(row=row_i, column=5, value=r["liters"])
        ws.cell(row=row_i, column=6, value=r["price_per_liter"])
        ws.cell(row=row_i, column=7, value=r["total_price"])
        ws.cell(row=row_i, column=8, value=currency)
        ws.cell(row=row_i, column=9, value=r["note"] or "")
        for col in range(1, 10):
            cell = ws.cell(row=row_i, column=col)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(horizontal="center")
            if col in (4, 5, 6, 7):
                cell.number_format = "#,##0.00"
        total_liters += r["liters"]
        total_price_bgn += to_bgn_by_date(r["total_price"], r["date"])

    total_row = len(rows) + 3
    ws.cell(row=total_row, column=2, value="ОБЩО (лв):")
    ws.cell(row=total_row, column=5, value=total_liters)
    ws.cell(row=total_row, column=7, value=total_price_bgn)
    for col in (2, 5, 7):
        cell = ws.cell(row=total_row, column=col)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        if col in (5, 7):
            cell.number_format = "#,##0.00"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    today = date.today().strftime("%d%m%Y")
    name = f"Подробен_отчет_{vehicle or 'автопарк'}_{today}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=sanitize_filename(name),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------- Резервно копие ----------------
@app.route("/backup")
@login_required
def backup():
    if not os.path.exists(DB_PATH):
        abort(404)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(DB_PATH, as_attachment=True, download_name=f"data_backup_{ts}.db")


# ---------------- Health check (за Railway) ----------------
@app.route("/healthz")
def healthz():
    return jsonify(status="ok")


init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
